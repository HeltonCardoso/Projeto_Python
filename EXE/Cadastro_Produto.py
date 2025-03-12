import os
import warnings
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.ttk import Progressbar
from datetime import datetime

# Suprimir o warning específico do openpyxl
warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

# Funções para copiar e reaplicar as validações de dados
from openpyxl.worksheet.datavalidation import DataValidation

def copiar_validacoes(worksheet):
    """ Copia todas as validações de dados da planilha """
    return list(worksheet.data_validations.dataValidation)

def reaplicar_validacoes(worksheet, validacoes):
    """ Reaplica as validações de dados na planilha """
    for dv in validacoes:
        worksheet.add_data_validation(dv)

class TelaCadastroProduto:
    def __init__(self, root):
        self.root = root
        self.root.title("CADASTRO PLANILHA ATHUS V1.0")
        self.centralizar_janela(750, 510)

        # Adicionar ícone à janela
        try:
            self.root.iconbitmap("icone.ico")  # Substitua "icone.ico" pelo caminho do seu ícone
        except:
            pass  # Ignora se o ícone não for encontrado

        # Configuração de estilo
        self.style = ttk.Style()
        self.style.configure("TButton", padding=6, font=("Helvetica", 10))
        self.style.configure("TLabel", font=("Helvetica", 10))
        self.style.configure("TEntry", font=("Helvetica", 10))

        # Interface
        frame = tk.Frame(self.root)
        frame.pack(pady=20)

        tk.Label(frame, text="SELECIONE PLANILHA ONLINE:").grid(row=0, column=0)
        self.entrada_origem = ttk.Entry(frame, width=70)
        self.entrada_origem.grid(row=0, column=1)
        ttk.Button(frame, text="Buscar", command=lambda: self.selecionar_arquivo(self.entrada_origem)).grid(row=0, column=2)

        tk.Label(frame, text="SELECIONE MODELO ATHUS:").grid(row=1, column=0)
        self.entrada_destino = ttk.Entry(frame, width=70)
        self.entrada_destino.grid(row=1, column=1)
        ttk.Button(frame, text="Buscar", command=lambda: self.selecionar_arquivo(self.entrada_destino)).grid(row=1, column=2)

        ttk.Button(self.root, text="IMPORTAR PLANILHA ONLINE", command=self.processar).pack(pady=10)

        # Barra de progresso
        self.progresso = Progressbar(self.root, orient="horizontal", length=500, mode="determinate")
        self.progresso.pack(pady=5)

        # Label de status
        self.status_label = tk.Label(self.root, text="Pronto para iniciar...", fg="blue")
        self.status_label.pack(pady=5)

        # Área de logs
        self.log_area = scrolledtext.ScrolledText(self.root, width=70, height=10, state="disabled")
        self.log_area.pack(pady=10)

        ttk.Button(self.root, text="FECHAR", command=self.root.destroy).pack(pady=10)

    # Rodapé
        frame_rodape = tk.Frame(root)
        frame_rodape.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Relógio no rodapé
        self.relogio = tk.Label(frame_rodape, font=("Arial", 10), fg="blue")
        self.relogio.pack(side=tk.RIGHT, padx=10)
        self.atualizar_relogio()
    
    def atualizar_relogio(self):
        agora = datetime.now()
        data_hora = agora.strftime("%d/%m/%Y %H:%M:%S")
        self.relogio.config(text=data_hora)
        self.root.after(1000, self.atualizar_relogio)  # Atualiza a cada 1 segundo

    def centralizar_janela(self, largura, altura):
        largura_tela = self.root.winfo_screenwidth()
        altura_tela = self.root.winfo_screenheight()
        pos_x = (largura_tela // 2) - (largura // 2)
        pos_y = (altura_tela // 2) - (altura // 2)
        self.root.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")

    def selecionar_arquivo(self, entrada):

        # Desativa o atributo -topmost temporariamente
        self.root.attributes('-topmost', False)
        
        # Abre o filedialog
        caminho = filedialog.askopenfilename(parent=self.root, filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
        
        # Reativa o atributo -topmost após o fechamento do diálogo
        self.root.attributes('-topmost', True)
        
        if caminho:
            entrada.delete(0, tk.END)
            entrada.insert(0, caminho)

    def log(self, mensagem):
        """Adiciona uma mensagem à área de logs."""
        self.log_area.config(state="normal")
        self.log_area.insert(tk.END, mensagem + "\n")
        self.log_area.config(state="disabled")
        self.log_area.yview(tk.END)  # Rola para a última linha

    def processar(self):
        planilha_origem = self.entrada_origem.get()
        planilha_destino = self.entrada_destino.get()

        if not planilha_origem or not planilha_destino:
            messagebox.showwarning("Atenção", "Selecione os arquivos de origem e destino!", parent=self.root)
            return

        try:
            caminho_arquivo_salvo = self.executar_processamento(planilha_origem, planilha_destino)
            messagebox.showinfo("Sucesso", f"Cadastro realizado com sucesso!\nArquivo salvo em:\n{caminho_arquivo_salvo}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento: {e}", parent=self.root)

    def executar_processamento(self, planilha_origem, planilha_destino):
        try:
            # Ler os dados da planilha de origem
            df = pd.read_excel(planilha_origem)
            total_linhas = len(df)
            self.progresso["maximum"] = total_linhas
            self.progresso["value"] = 0

            # Criar listas de DataFrames para cada aba
            dados_sheets = {
                "PRODUTO": [],
                "PRECO": [],
                "LOJA WEB": [],
                "KIT": [],
                "VOLUME": []
            }

            # Criar dicionário {EAN: Nome do Produto}
            produto_dict = {
                str(row["EAN"]).strip(): row["NOMEONCLICK"] if pd.notna(row["NOMEONCLICK"]) else "Nome Desconhecido"
                for _, row in df.iterrows()
            }

            # Criar conjunto para armazenar a marca cadastrada
            marcas_cadastradas = set()
            data_atual = datetime.now()

            # Formatar a data atual no formato desejado (DD/MM/YYYY)
            data_formatada = data_atual.strftime("%d/%m/%Y")

            # Adicionar 30 anos à data atual
            data_mais_20_anos = data_atual.replace(year=data_atual.year + 30)

            # Formatar a nova data no mesmo formato
            data_formatada_mais_20_anos = data_mais_20_anos.strftime("%d/%m/%Y")

            for idx, row in df.iterrows():
                ean = str(row["EAN"]).strip()
                tipo_produto = row["TIPODEPRODUTO"].strip().upper()
                descricao = row["PRODUTO"]
                quantidade = row["QTDECOMPONENTES"]
                cor = row["COR"]
                altura = row["EMBALTURA"]
                largura = row["EMBLARGURA"]
                comprimento = row["EMBCOMPRIMENTO"]
                volumes = int(row["VOLUMES"]) if pd.notna(row["VOLUMES"]) else 1  # Garante que tenha pelo menos 1 volume
                componentes = row["EANCOMPONENTES"]
                marca = row["MARCA"]
                custo = row["CUSTO"]
                preco_venda = row["DE"]
                preco_promo = row["POR"]
                fornecedor = row["FORNECEDOR"]
                outros = row["OUTROS"]
                ncm = row["NCM"]
                cod_forn = row["CODFORN"]
                nome_onclick = row["NOMEONCLICK"]
                categoria = row["CATEGORIA"]
                grupo = row["GRUPO"]
                nome_ecommerce = row["NOMEE-COMMERCE"]
                disponibilidade_web = row["DISPONIBILIDADEWEB"]
                descricao_html = row["DESCRICAOHTML"]
                peso_bruto = row["PESOBRUTO"]
                peso_liquido = row["PESOLIQUIDO"]
                vol_peso_bruto = row["VOLPESOBRUTO"]
                vol_peso_liquido = row["VOLPESOLIQ"]
                vol_largura = row["VOLLARGURA"]
                vol_altura = row["VOLALTURA"]
                vol_comprimento = row["VOLCOMPRIMENTO"]

                tipo_produto_valor = 0 if tipo_produto == "PRODUTO ACABADO" else 2

                # Regra: Nome do produto (5ª coluna) com apenas 25 caracteres
                nome_reduzido = nome_onclick[:25] if isinstance(nome_onclick, str) else ""

                # Adicionar marca ao conjunto
                marcas_cadastradas.add(marca)

                # Usar sempre as medidas PESOBRUTO, PESOLIQUIDO, EMBLARGURA, EMBALTURA, EMBCOMPRIMENTO para a aba "PRODUTO"
                peso_bruto_final = peso_bruto
                peso_liquido_final = peso_liquido
                largura_final = largura
                altura_final = altura
                comprimento_final = comprimento

                # Adicionar dados ao dicionário de listas
                dados_sheets["PRODUTO"].append([
                    ean, cod_forn, tipo_produto_valor, nome_onclick, nome_reduzido, nome_onclick, nome_onclick, "",
                    marca, categoria, grupo, "", "", nome_ecommerce, "", "", "F", "F", "F", "", volumes,
                    peso_bruto_final, peso_liquido_final, largura_final, altura_final, comprimento_final, "", 90, 1000,
                    disponibilidade_web, "F", "F", ncm, "", "0", "T", "F", "F", "NAO", nome_ecommerce, marca,
                    "90 dias após o recebimento do produto", disponibilidade_web, descricao_html, "F", "F"
                ])

                if tipo_produto == "KIT" and pd.notna(componentes):
                    for comp in str(componentes).split("/"):
                        comp_ean = comp.strip()
                        nome_componente = produto_dict.get(comp_ean, "Desconhecido")
                        dados_sheets["KIT"].append([ean, comp_ean, nome_componente, "1", "", "0"])

                # Adicionar volumes (Agora inclui produtos com apenas 1 volume)
                for i in range(volumes):
                    if volumes == 1:
                        # Quando for 1 volume, usar dados da aba "PRODUTO"
                        dados_sheets["VOLUME"].append([
                            ean, nome_onclick, peso_bruto_final, peso_liquido_final, largura_final, altura_final, "",
                            comprimento_final, "", "BOX", "T", i + 1
                        ])
                    else:
                        # Quando houver mais de 1 volume, usar os dados de volume específico
                        dados_sheets["VOLUME"].append([
                            ean, nome_onclick, vol_peso_bruto, vol_peso_liquido, vol_largura, vol_altura, "",
                            vol_comprimento, "", "BOX", "T", i + 1
                        ])

                dados_sheets["PRECO"].append([
                    ean, fornecedor, custo, outros, "", "", row["CUSTOTOTAL"], preco_venda, preco_promo, preco_promo,
                    data_formatada, data_formatada_mais_20_anos, "", "F"
                ])

                dados_sheets["LOJA WEB"].append([
                    ean, "", "", "", row["CATEGORIAPRINCIPALTRAY"], "", "", "", "T", "T", "", "", "",
                    row["CATEGORIAPRINCIPALCORP"], row["NIVELADICIONAL1CORP"], "", "", "T", "T"
                ])

                # Atualizar a barra de progresso e o status
                self.progresso["value"] = idx + 1
                self.status_label.config(text=f"Processando linha {idx + 1} de {total_linhas}...")
                self.log(f"Processando produto: {nome_onclick}")
                self.root.update_idletasks()

            # Carregar a planilha original mantendo formatação
            wb = load_workbook(planilha_destino)

            # Obter a aba "Tipo Importacao" e copiar as validações antes de salvar
            ws_tipo_importacao = wb["Tipo Importacao"]
            validacoes_tipo_importacao = copiar_validacoes(ws_tipo_importacao)

            # Sobrescrever os dados mantendo formatação
            for sheet_name, data in dados_sheets.items():
                ws = wb[sheet_name]

                # Definir linha inicial dependendo da aba
                if sheet_name == "PRODUTO":
                    start_row = 3  # Para a aba PRODUTO, começa na linha 3
                else:
                    start_row = 2  # Para as outras abas, começa na linha 2

                # Limpar apenas os dados (não os cabeçalhos)
                for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None

                # Escrever os novos dados
                for i, row_data in enumerate(data, start=start_row):
                    for j, value in enumerate(row_data, start=1):
                        ws.cell(row=i, column=j, value=value)

            # Gerar o nome do novo arquivo com a marca
            if marcas_cadastradas:
                marca_unica = next(iter(marcas_cadastradas))  # Obtém a única marca do conjunto
                novo_nome_arquivo = f"Template_Produtos_Mpozenato_CADASTRO_{marca_unica}.xlsx"

                # Caminho relativo para a pasta PLANILHA_PREENCHIDA na raiz do projeto
                pasta_destino = os.path.join(os.getcwd(), "PLANILHA_PREENCHIDA")

                # Verificar se o diretório existe, caso contrário, criar
                if not os.path.exists(pasta_destino):
                    os.makedirs(pasta_destino)

                caminho_arquivo = os.path.join(pasta_destino, novo_nome_arquivo)  # Combina o diretório e o nome do arquivo

                try:
                    wb.save(caminho_arquivo)
                    self.log("Arquivo salvo com sucesso!")
                except PermissionError:
                    messagebox.showerror("Erro", "Feche o arquivo de saída e tente novamente!", parent=self.root)
                    return

                # Reabrir a planilha e reaplicar as validações na aba "Tipo Importação"
                wb = load_workbook(caminho_arquivo)
                ws_tipo_importacao = wb["Tipo Importacao"]
                reaplicar_validacoes(ws_tipo_importacao, validacoes_tipo_importacao)

                # Salvar novamente após reaplicar as validações
                wb.save(caminho_arquivo)

            return caminho_arquivo  # Retorna o caminho completo do arquivo salvo

        except Exception as e:
            self.log(f"Erro durante o processamento: {e}", parent=self.root)
            raise e

# Função principal
if __name__ == "__main__":
    root = tk.Tk()
    app = TelaCadastroProduto(root)
    root.mainloop()